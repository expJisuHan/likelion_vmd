from __future__ import annotations

import zipfile
from datetime import datetime
from typing import Any
from xml.sax.saxutils import escape

from ..config import settings
from ..services.records import excel_column_widths_for_zone, excel_headers_for_zone, record_zone, result_to_row
from ..utils import image_data_url_to_media, image_data_url_to_thumbnail, timestamp

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - pure-Python XLSX fallback remains available.
    Workbook = None
    ExcelImage = None


def excel_column_letter(col_idx: int) -> str:
    letters = ""
    while col_idx:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def xlsx_inline_cell(row_idx: int, col_idx: int, value: Any, style: int | None = None) -> str:
    ref = f"{excel_column_letter(col_idx)}{row_idx}"
    style_attr = f' s="{style}"' if style is not None else ""
    raw_text = "" if value is None else str(value)
    cleaned = "".join(char for char in raw_text if char in {"\t", "\n", "\r"} or ord(char) >= 0x20)
    text = escape(cleaned)
    return f'<c r="{ref}" t="inlineStr"{style_attr}><is><t xml:space="preserve">{text}</t></is></c>'


def save_excel_pure_python(records: list[dict[str, Any]], prefix: str) -> tuple[str, bytes]:
    """openpyxl이 없는 환경을 위한 내장 XLSX 생성기(zip + 수기 OOXML)."""
    path = settings.excel_dir / f"{prefix}_{timestamp()}.xlsx"
    zone = record_zone(records[0]) if records else "VP"
    headers = excel_headers_for_zone(zone)
    rows = [headers, *[result_to_row(record, zone) for record in records]]
    image_parts = []
    for row_idx, record in enumerate(records, start=2):
        media = image_data_url_to_media(record.get("image_data_url", ""))
        if media is None:
            continue
        image_parts.append(
            {
                "row_idx": row_idx,
                "data": media["data"],
                "extension": media["extension"],
                "content_type": media["content_type"],
                "rotation": media["rotation"],
                "cx": media["cx"],
                "cy": media["cy"],
                "name": f'image{len(image_parts) + 1}.{media["extension"]}',
                "rid": f"rId{len(image_parts) + 1}",
            }
        )

    content_types = [
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
        '<Default Extension="xml" ContentType="application/xml"/>',
    ]
    for extension in sorted({part["extension"] for part in image_parts}):
        content_type = "image/png" if extension == "png" else "image/jpeg"
        content_types.append(f'<Default Extension="{extension}" ContentType="{content_type}"/>')
    content_types.extend(
        [
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>',
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>',
            '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>',
            '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>',
        ]
    )
    if image_parts:
        content_types.append('<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')

    col_widths = excel_column_widths_for_zone(zone)
    cols_xml = "".join(f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>' for idx, width in col_widths.items())
    row_xml = []
    for row_idx, values in enumerate(rows, start=1):
        attrs = f' r="{row_idx}"'
        if row_idx == 1:
            attrs += ' ht="24" customHeight="1"'
        if row_idx > 1:
            attrs += ' ht="72" customHeight="1"'
        cells = "".join(xlsx_inline_cell(row_idx, col_idx, value, 1 if row_idx == 1 else None) for col_idx, value in enumerate(values, start=1))
        row_xml.append(f"<row{attrs}>{cells}</row>")
    drawing_ref = '<drawing r:id="rId1"/>' if image_parts else ""
    worksheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
        "</sheetView></sheetViews>"
        f"<cols>{cols_xml}</cols>"
        f'<sheetData>{"".join(row_xml)}</sheetData>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>'
        f"{drawing_ref}"
        "</worksheet>"
    )

    drawing_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
    )
    for idx, part in enumerate(image_parts, start=1):
        zero_based_row = part["row_idx"] - 1
        drawing_xml += (
            "<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>0</xdr:col><xdr:colOff>95250</xdr:colOff><xdr:row>{zero_based_row}</xdr:row><xdr:rowOff>95250</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{part["cx"]}" cy="{part["cy"]}"/>'
            "<xdr:pic>"
            f'<xdr:nvPicPr><xdr:cNvPr id="{idx}" name="Upload Image {idx}"/><xdr:cNvPicPr/></xdr:nvPicPr>'
            f'<xdr:blipFill><a:blip r:embed="{part["rid"]}"/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
            f'<xdr:spPr><a:xfrm rot="{part["rotation"]}"><a:off x="0" y="0"/><a:ext cx="{part["cx"]}" cy="{part["cy"]}"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
            "</xdr:pic><xdr:clientData/>"
            "</xdr:oneCellAnchor>"
        )
    drawing_xml += "</xdr:wsDr>"

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as xlsx:
        xlsx.writestr("[Content_Types].xml", f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">{"".join(content_types)}</Types>')
        xlsx.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
            '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
            "</Relationships>",
        )
        xlsx.writestr(
            "docProps/core.xml",
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><dc:title>VMD 분석 결과</dc:title><dc:creator>AX R&amp;D VMD</dc:creator><dcterms:created xsi:type="dcterms:W3CDTF">{datetime.utcnow().isoformat()}Z</dcterms:created></cp:coreProperties>',
        )
        xlsx.writestr(
            "docProps/app.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"><Application>AX R&amp;D VMD</Application></Properties>',
        )
        xlsx.writestr(
            "xl/workbook.xml",
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="{zone} 결과" sheetId="1" r:id="rId1"/></sheets></workbook>',
        )
        xlsx.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>',
        )
        xlsx.writestr(
            "xl/styles.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><fonts count="2"><font><sz val="11"/><color theme="1"/><name val="Calibri"/></font><font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font></fonts><fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FF1F4E78"/><bgColor indexed="64"/></patternFill></fill></fills><borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders><cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs><cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf><xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf></cellXfs><cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles></styleSheet>',
        )
        xlsx.writestr("xl/worksheets/sheet1.xml", worksheet_xml)
        if image_parts:
            xlsx.writestr(
                "xl/worksheets/_rels/sheet1.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>',
            )
            drawing_rels = "".join(
                f'<Relationship Id="{part["rid"]}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/{part["name"]}"/>'
                for part in image_parts
            )
            xlsx.writestr(
                "xl/drawings/_rels/drawing1.xml.rels",
                f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{drawing_rels}</Relationships>',
            )
            xlsx.writestr("xl/drawings/drawing1.xml", drawing_xml)
            for part in image_parts:
                xlsx.writestr(f'xl/media/{part["name"]}', part["data"])
    return str(path.relative_to(settings.project_root)), path.read_bytes()


def save_excel(records: list[dict[str, Any]], prefix: str = "vmd_results") -> tuple[str, bytes]:
    settings.ensure_dirs()
    if Workbook is None:
        return save_excel_pure_python(records, prefix)

    zone = record_zone(records[0]) if records else "VP"
    headers = excel_headers_for_zone(zone)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{zone} 결과"
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    image_streams = []
    for record in records:
        row_idx = ws.max_row + 1
        ws.append(result_to_row(record, zone))
        if ExcelImage is not None:
            stream = image_data_url_to_thumbnail(record.get("image_data_url", ""))
            if stream is not None:
                image_streams.append(stream)
                embedded = ExcelImage(stream)
                embedded.anchor = f"A{row_idx}"
                ws.add_image(embedded)
                ws.row_dimensions[row_idx].height = 72

    widths = excel_column_widths_for_zone(zone)
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(col_idx, 14)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    path = settings.excel_dir / f"{prefix}_{timestamp()}.xlsx"
    wb.save(path)
    return str(path.relative_to(settings.project_root)), path.read_bytes()
