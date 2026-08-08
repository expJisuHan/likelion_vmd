from __future__ import annotations

import base64
import html
import json
import os
import re
import struct
import sys
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - pure-Python XLSX fallback remains available.
    Workbook = None

try:
    from openpyxl.drawing.image import Image as ExcelImage
    from PIL import Image as PillowImage
    from PIL import ImageOps
except Exception:  # pragma: no cover - image embedding is optional.
    ExcelImage = None
    PillowImage = None
    ImageOps = None

try:
    from reportlab.lib import colors as pdf_colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Image as ReportLabImage
    from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except Exception:  # pragma: no cover - PDF export reports a dependency error when unavailable.
    REPORTLAB_AVAILABLE = False


ROOT = Path(__file__).resolve().parents[1]
STATIC_DIR = ROOT / "app" / "static"
OUTPUT_DIR = ROOT / "outputs"
JSON_DIR = OUTPUT_DIR / "json"
EXCEL_DIR = OUTPUT_DIR / "excel"
PDF_DIR = OUTPUT_DIR / "pdf"

LMSTUDIO_BASE_URL = os.environ.get("LMSTUDIO_BASE_URL", "http://127.0.0.1:1234/v1").rstrip("/")
DEFAULT_MODEL = os.environ.get("LMSTUDIO_MODEL", "google/gemma-4-12b")
DEFAULT_FALLBACK_MODELS = ["google/gemma-4-12b", "google/gemma-4-12b-qat", "google/gemma-3-4b"]


def ensure_dirs() -> None:
    JSON_DIR.mkdir(parents=True, exist_ok=True)
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    PDF_DIR.mkdir(parents=True, exist_ok=True)


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def safe_file_name(name: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z가-힣_. -]+", "_", name).strip()
    return cleaned or "image"


def normalize_zone(value: str | None) -> str:
    value = (value or "VP").upper().strip()
    return value if value in {"VP", "PP", "IP"} else "VP"


def grade_from_score(score: Any) -> str:
    try:
        numeric = float(score)
    except (TypeError, ValueError):
        return "평가 보류"
    if numeric >= 90:
        return "매우 우수"
    if numeric >= 80:
        return "우수"
    if numeric >= 70:
        return "보통"
    if numeric >= 60:
        return "개선 필요"
    return "집중 개선"


ZONE_CRITERIA: dict[str, list[str]] = {
    "VP": [
        "연출 콘셉트의 시각적 일관성",
        "시각적 핵심 상품의 가시성과 강조도",
        "포컬 포인트 및 시선 유도",
        "구성·비례·여백의 균형",
        "색채·조명·소품의 조화",
    ],
    "IP": [
        "상품 분류 및 구획의 명확성",
        "컬러 배열의 규칙성과 연속성",
        "사이즈 배열의 정확성",
        "상품 비교 및 선택 편의성",
        "진열 수량과 공간 점유의 적절성",
        "행거 간 간격 및 페이싱·진열 정돈 상태",
    ],
    "PP": [
        "핵심 제안 상품의 차별화와 위계",
        "상품 그룹화 및 연계 제안의 명확성",
        "페이스아웃과 상품 특징 노출",
        "진열 밀도와 공간 점유의 적절성",
        "상품 간격·방향·배열선의 정돈",
        "POP·가격·프로모션 정보의 연결성과 가독성",
    ],
}


EXCEL_COMMON_HEADERS = [
    "사진",
    "지정 구역",
    "모델 감지 구역",
    "총점",
    "사진 품질 점수",
    "사진 품질 코멘트",
    "영역별 평가",
    "최종 요약 답변",
    "감지된 문제점",
    "개선 방향",
    "전체 개선 요약",
    "분석 시간",
]


def criteria_for_zone(zone: str | None) -> list[str]:
    return ZONE_CRITERIA[normalize_zone(zone)]


def excel_headers_for_zone(zone: str | None) -> list[str]:
    headers = list(EXCEL_COMMON_HEADERS)
    for criterion in criteria_for_zone(zone):
        headers.extend(
            [
                f"{criterion} 점수",
                f"{criterion} 근거",
                f"{criterion} 문제점",
                f"{criterion} 개선안",
            ]
        )
    return headers


def excel_column_widths_for_zone(zone: str | None) -> dict[int, float]:
    if normalize_zone(zone) == "VP":
        values = [18, 14, 14, 14, 14, 22, 14, 22, 14, 42, 42, 14, 42, 42, 22, 22, 22, 22, 22, 14, 36, 36, 36, 22, 14, 36, 36, 36, 22, 14, 36, 36]
    else:
        values = [18, 14, 14, 22, 14, 42, 42, 42, 42, 42, 42, 14]
        for _criterion in criteria_for_zone(zone):
            values.extend([14, 36, 36, 36])
    return {index: width for index, width in enumerate(values, start=1)}


def format_elapsed(seconds: float | int | None) -> str:
    if seconds is None:
        return ""
    total_seconds = max(0, int(round(float(seconds))))
    minutes, remaining_seconds = divmod(total_seconds, 60)
    if minutes:
        return f"{minutes}분 {remaining_seconds}초"
    return f"{remaining_seconds}초"


def vmd_json_schema() -> dict[str, Any]:
    score_props = {
        "layout": {"type": "integer", "minimum": 0, "maximum": 100},
        "presentation_mood": {"type": "integer", "minimum": 0, "maximum": 100},
        "brand_fit": {"type": "integer", "minimum": 0, "maximum": 100},
        "color_harmony": {"type": "integer", "minimum": 0, "maximum": 100},
        "cleanliness": {"type": "integer", "minimum": 0, "maximum": 100},
        "customer_attention": {"type": "integer", "minimum": 0, "maximum": 100},
        "season_concept_fit": {"type": "integer", "minimum": 0, "maximum": 100},
    }
    return {
        "type": "object",
        "properties": {
            "user_selected_zone": {"type": "string", "enum": ["VP", "PP", "IP"]},
            "ai_detected_zone": {"type": "string", "enum": ["VP", "PP", "IP", "UNKNOWN"]},
            "zone_confidence": {"type": "number", "minimum": 0, "maximum": 1},
            "store_type_assumption": {"type": "string"},
            "photo_quality": {
                "type": "object",
                "properties": {
                    "score": {"type": "integer", "minimum": 0, "maximum": 100},
                    "is_blurry": {"type": "boolean"},
                    "is_tilted": {"type": "boolean"},
                    "has_background_interference": {"type": "boolean"},
                    "needs_retake": {"type": "boolean"},
                    "comment": {
                        "type": "string",
                        "description": "촬영 상태와 평가 신뢰도에 미치는 영향을 근거와 함께 설명한 2~3문장",
                    },
                },
                "required": [
                    "score",
                    "is_blurry",
                    "is_tilted",
                    "has_background_interference",
                    "needs_retake",
                    "comment",
                ],
                "additionalProperties": False,
            },
            "mannequin": {
                "type": "object",
                "properties": {
                    "exists": {"type": "boolean"},
                    "type": {"type": "string"},
                    "has_head": {"type": "boolean"},
                    "comment": {
                        "type": "string",
                        "description": "마네킹이 있으면 판정 근거와 연출 상태를 1~2문장으로 설명하고, 없으면 빈 문자열",
                    },
                },
                "required": ["exists", "type", "has_head", "comment"],
                "additionalProperties": False,
            },
            "obstacles": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "object": {"type": "string"},
                        "location": {"type": "string"},
                        "severity": {"type": "string", "enum": ["low", "medium", "high", "unknown"]},
                        "reason": {
                            "type": "string",
                            "description": "해당 물체가 존의 목적, 상품 시야 또는 고객 동선에 미치는 영향을 1~2문장으로 설명",
                        },
                    },
                    "required": ["object", "location", "severity", "reason"],
                    "additionalProperties": False,
                },
            },
            "scores": {
                "type": "object",
                "properties": score_props,
                "required": list(score_props.keys()),
                "additionalProperties": False,
            },
            "total_score": {"type": "integer", "minimum": 0, "maximum": 100},
            "grade": {"type": "string"},
            "positive_points": {
                "type": "array",
                "description": "사진에서 확인 가능한 근거를 포함한 강점 2~3개",
                "items": {"type": "string"},
                "minItems": 2,
                "maxItems": 3,
            },
            "critical_issues": {
                "type": "array",
                "description": "문제의 위치, 관찰 근거, VMD 영향을 포함한 핵심 문제 3~4개",
                "items": {"type": "string"},
                "minItems": 3,
                "maxItems": 4,
            },
            "improvement_suggestions": {
                "type": "array",
                "description": "담당자가 바로 실행할 수 있도록 대상, 방법, 기대 효과를 포함한 개선안 3~4개",
                "items": {"type": "string"},
                "minItems": 3,
                "maxItems": 4,
            },
            "final_summary": {
                "type": "string",
                "description": "현재 상태, 핵심 강점과 문제, 우선 조치 순서를 종합한 3~5문장",
            },
            "zone_evaluation_summary": {
                "type": "string",
                "description": "사용자가 지정한 VP/IP/PP 영역 기준으로 본 종합 평가. reference Excel의 영역별 평가 컬럼에 들어갈 내용",
            },
            "priority_action_summary": {
                "type": "string",
                "description": "가장 먼저 조치할 사항과 이유를 설명한 최종 요약 답변",
            },
            "detected_issues": {
                "type": "array",
                "description": "감지된 문제점을 줄 단위로 정리",
                "items": {"type": "string"},
            },
            "improvement_actions": {
                "type": "array",
                "description": "개선 방향을 줄 단위로 정리",
                "items": {"type": "string"},
            },
            "overall_improvement_summary": {
                "type": "string",
                "description": "전체 개선 요약. improvement_actions를 종합해 한 문단 또는 줄바꿈 목록으로 작성",
            },
            "criteria_evaluations": {
                "type": "array",
                "description": "사용자 지정 존의 평가항목별 점수, 근거, 문제점, 개선안",
                "items": {
                    "type": "object",
                    "properties": {
                        "criterion": {"type": "string"},
                        "score": {"type": ["integer", "null"], "minimum": 0, "maximum": 100},
                        "evidence": {"type": "string"},
                        "issue": {"type": "string"},
                        "suggestion": {"type": "string"},
                    },
                    "required": ["criterion", "score", "evidence", "issue", "suggestion"],
                    "additionalProperties": False,
                },
            },
        },
        "required": [
            "user_selected_zone",
            "ai_detected_zone",
            "zone_confidence",
            "store_type_assumption",
            "photo_quality",
            "mannequin",
            "obstacles",
            "scores",
            "total_score",
            "grade",
            "positive_points",
            "critical_issues",
            "improvement_suggestions",
            "final_summary",
            "zone_evaluation_summary",
            "priority_action_summary",
            "detected_issues",
            "improvement_actions",
            "overall_improvement_summary",
            "criteria_evaluations",
        ],
        "additionalProperties": False,
    }


def system_prompt() -> str:
    return (
        "당신은 백화점과 패션 리테일 매장을 평가하는 Visual Merchandising(VMD) 전문가입니다. "
        "IP, PP, VP 존의 역할을 이해하고, 사진 속 매장 연출을 전문가 관점에서 평가합니다. "
        "답변은 반드시 JSON 스키마에 맞는 JSON 객체 하나만 반환합니다. 마크다운, 설명문, 코드블록은 금지합니다. "
        "평가는 칭찬 위주가 아니라 실제 개선 가능한 문제를 찾는 비판적 분석이어야 합니다. "
        "다만 표현은 사용자가 받아들이기 쉬운 부드러운 전문가 톤을 유지합니다. "
        "마네킹이 없는 사진에서는 마네킹 코멘트를 작성하지 말고, 먼저 마네킹 유무를 판단하세요. "
        "머리 있는 마네킹, 머리 없는 바디 마네킹, 옷걸이/상하의 진열을 구분하세요. "
        "의자, 상자, 공기청정기, 화분, 테이블, 적재물 등은 상품 시야나 동선을 방해하면 방해물로 기록하세요. "
        "구조물, 소품, 상품, 방해물을 혼동하지 마세요. "
        "사진의 선명도, 조명, 질감, 각도, 기울어짐, 배경 간섭, 핵심 존 가시성을 별도 평가하세요. "
        "사용자가 VP/PP/IP 존을 지정하면 그 존의 목적에 맞춰 평가하고, AI 판단 존과 신뢰도는 별도로 반환하세요. "
        "평가 문장은 사진에서 실제로 관찰한 대상과 위치를 근거로 작성하고, 점수만 되풀이하거나 한 문장으로 뭉뚱그리지 마세요. "
        "각 강점과 문제점은 관찰 사실과 VMD 영향을 포함하고, 각 개선안은 무엇을 어떻게 바꿀지와 기대 효과까지 설명하세요."
    )


def build_user_text(options: dict[str, Any], image_count: int) -> str:
    zone = normalize_zone(options.get("zoneMode"))
    store_type = options.get("storeType", "UNKNOWN")
    tone = options.get("tone", "SOFT_CRITICAL")
    criteria = options.get("criteria", [])
    focus_keywords = options.get("focusKeywords", [])
    extra = (options.get("extraCriteria") or "").strip()
    if not isinstance(focus_keywords, list):
        focus_keywords = []
    focus_keywords = [str(keyword).strip() for keyword in focus_keywords if str(keyword).strip()]

    zone_instruction = (
        f"사용자가 이 이미지를 {zone} 존으로 지정했습니다. "
        f"AI 판단 존도 반환하되, 평가는 {zone} 존 기준으로 하세요."
    )
    zone_criteria = criteria_for_zone(zone)
    return "\n".join(
        [
            f"이미지 {image_count}장을 함께 보고 하나의 VMD 평가 결과를 작성하세요.",
            zone_instruction,
            f"사용자 선택 존: {zone}",
            f"매장 유형 옵션: {store_type}",
            f"분석 톤 옵션: {tone}",
            "기본 평가 항목: " + (", ".join(criteria) if criteria else "전체 기본 항목"),
            f"{zone} 존 전용 평가 항목: " + ", ".join(zone_criteria),
            "참고 키워드: " + (", ".join(focus_keywords) if focus_keywords else "없음"),
            "추가 평가 요청: " + (extra if extra else "없음"),
            "결과 분량 기준을 반드시 지키세요.",
            "- zone_evaluation_summary: 지정 존의 목적과 이미지 관찰 근거를 연결해 3~5문장으로 작성하세요.",
            "- priority_action_summary: 가장 먼저 해야 할 조치와 이유를 3~5문장으로 작성하세요.",
            "- detected_issues: 감지된 문제점을 0~4개 작성하세요. 문제가 거의 없으면 빈 배열로 두세요.",
            "- improvement_actions: 실행 가능한 개선 방향을 2~4개 작성하세요.",
            "- overall_improvement_summary: 개선 방향 전체를 한 문단 또는 줄바꿈 목록으로 요약하세요.",
            "- criteria_evaluations: 아래 존 전용 평가 항목을 같은 이름과 같은 순서로 모두 작성하세요.",
            *[f"  {index}. {criterion}" for index, criterion in enumerate(zone_criteria, start=1)],
            "- positive_points: 서로 다른 강점 2~3개. 각 항목은 관찰 근거와 효과를 담은 1~2문장으로 작성하세요.",
            "- critical_issues: 서로 다른 핵심 문제 3~4개. 각 항목은 위치/대상, 관찰 근거, VMD 영향을 담은 2문장 내외로 작성하세요.",
            "- improvement_suggestions: 문제점에 대응하는 실행안 3~4개. 각 항목은 수정 대상, 구체적인 방법, 기대 효과를 담은 2문장 내외로 작성하세요.",
            "- final_summary: 현재 상태와 우선순위를 종합한 3~5문장으로 작성하세요. 목록 내용을 그대로 반복하지 마세요.",
            "- photo_quality.comment: 촬영 상태와 분석 신뢰도 영향을 2~3문장으로 작성하세요.",
            "사진에서 확인할 수 없는 사실을 분량을 채우기 위해 추측하지 마세요.",
            "모든 점수는 0~100 정수로 작성하세요.",
        ]
    )


def lmstudio_request(payload: dict[str, Any], timeout: int = 180) -> dict[str, Any]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        f"{LMSTUDIO_BASE_URL}/chat/completions",
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        detail = body.strip() or exc.reason
        raise RuntimeError(f"LM Studio HTTP {exc.code}: {detail}") from exc


def lmstudio_model_candidates(requested_model: str) -> list[str]:
    env_fallbacks = os.environ.get("LMSTUDIO_FALLBACK_MODELS", "")
    fallback_models = [item.strip() for item in env_fallbacks.split(",") if item.strip()]
    fallback_models.extend(DEFAULT_FALLBACK_MODELS)
    candidates: list[str] = []
    for model in [requested_model, *fallback_models]:
        if model and model not in candidates:
            candidates.append(model)
    return candidates


def is_retriable_lmstudio_error(message: str) -> bool:
    retriable_markers = [
        "Model unloaded",
        "Channel Error",
        "Failed to load model",
        "exited before becoming healthy",
        "vision",
        "image",
    ]
    return any(marker.lower() in message.lower() for marker in retriable_markers)


def check_lmstudio() -> tuple[bool, str]:
    try:
        request = urllib.request.Request(f"{LMSTUDIO_BASE_URL}/models", method="GET")
        with urllib.request.urlopen(request, timeout=3) as response:
            data = json.loads(response.read().decode("utf-8"))
        count = len(data.get("data", [])) if isinstance(data, dict) else 0
        return True, f"LM Studio connected ({count} model entries)."
    except Exception as exc:
        return False, str(exc)


def image_content_items(images: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for image in images:
        data_url = image.get("dataUrl", "")
        if not data_url.startswith("data:image/"):
            raise ValueError(f"Invalid image data for {image.get('name', 'image')}")
        items.append({"type": "image_url", "image_url": {"url": data_url}})
    return items


def parse_model_content(content: str) -> dict[str, Any]:
    cleaned = content.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned, flags=re.IGNORECASE).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            return json.loads(cleaned[start : end + 1])
        raise


def schema_instruction() -> str:
    example = {
        "user_selected_zone": "VP",
        "ai_detected_zone": "VP",
        "zone_confidence": 0.85,
        "store_type_assumption": "UNKNOWN",
        "photo_quality": {
            "score": 80,
            "is_blurry": False,
            "is_tilted": False,
            "has_background_interference": False,
            "needs_retake": False,
            "comment": (
                "사진은 핵심 연출과 주요 상품의 형태를 식별할 수 있을 만큼 선명합니다. "
                "다만 우측 하단의 주변 집기와 약간 기울어진 촬영 각도가 시선 흐름 판단을 일부 방해하므로, "
                "정면에서 수평을 맞춘 사진을 추가하면 평가 신뢰도가 높아집니다."
            ),
        },
        "mannequin": {
            "exists": True,
            "type": "headless_body_mannequin",
            "has_head": False,
            "comment": (
                "머리가 없는 인체형 바디와 입체적으로 착장된 의상이 중앙 전면에 배치되어 있어 "
                "헤드리스 마네킹으로 판단됩니다. 주력 상품의 실루엣과 스타일링을 전달하는 역할도 분명합니다."
            ),
        },
        "obstacles": [
            {
                "object": "chair",
                "location": "right bottom",
                "severity": "medium",
                "reason": (
                    "우측 하단 의자가 주력 상품으로 향하는 시선을 분산시키고 전면 접근 폭을 좁힙니다. "
                    "연출 소품으로 보기에는 주변 상품과의 연결성이 약해 방해물에 가깝습니다."
                ),
            }
        ],
        "scores": {
            "layout": 80,
            "presentation_mood": 80,
            "brand_fit": 80,
            "color_harmony": 80,
            "cleanliness": 80,
            "customer_attention": 80,
            "season_concept_fit": 80,
        },
        "total_score": 80,
        "grade": "우수",
        "positive_points": [
            (
                "중앙의 주력 착장이 주변 상품보다 높은 위치와 밝은 조명으로 강조되어 첫 시선이 자연스럽게 모입니다. "
                "VP 존의 핵심 역할인 대표 상품과 분위기 전달이 비교적 명확합니다."
            ),
            (
                "상의와 하의의 색상 계열이 반복되어 연출 전체에 통일감이 있습니다. "
                "고객이 코디 조합을 빠르게 이해하는 데 도움이 됩니다."
            ),
        ],
        "critical_issues": [
            (
                "중앙 마네킹 주변에 소품과 상품이 비슷한 간격으로 놓여 있어 주력 상품과 보조 상품의 위계가 약합니다. "
                "이 때문에 시선이 여러 곳으로 분산되고 핵심 시즌 메시지가 즉시 읽히지 않습니다."
            ),
            (
                "우측 하단의 의자가 전면 시야를 일부 가리고 상품 접근 폭을 좁힙니다. "
                "연출 의도가 분명하지 않아 VP 존의 집중도와 정돈감을 떨어뜨립니다."
            ),
            (
                "배경 상품의 색상 대비가 중앙 착장과 유사해 실루엣 경계가 흐려집니다. "
                "대표 룩이 배경에서 충분히 분리되지 않아 원거리 주목도가 낮아질 수 있습니다."
            ),
        ],
        "improvement_suggestions": [
            (
                "중앙 착장을 기준으로 보조 상품의 높이와 간격을 단계적으로 조정해 시각적 위계를 만드세요. "
                "대표 상품이 먼저 보이고 이후 연관 상품으로 시선이 이동하는 흐름을 만들 수 있습니다."
            ),
            (
                "우측 하단 의자를 제거하거나 연출 콘셉트와 연결되는 위치로 옮겨 전면 접근 폭을 확보하세요. "
                "상품 가시성과 고객 동선을 동시에 개선할 수 있습니다."
            ),
            (
                "중앙 착장 뒤 배경의 색상 수를 줄이고 명도 대비가 생기는 상품으로 교체하세요. "
                "주력 실루엣이 선명해져 원거리에서도 시즌 룩을 빠르게 인지할 수 있습니다."
            ),
        ],
        "final_summary": (
            "현재 VP 존은 중앙 착장과 색상 통일감 덕분에 기본적인 시즌 분위기는 전달하고 있습니다. "
            "다만 주변 상품의 위계가 약하고 우측 집기가 시선을 분산시켜 대표 룩의 집중도가 충분하지 않습니다. "
            "가장 먼저 전면 방해물을 정리하고 중앙 착장 주변의 간격과 높이를 재구성하는 것이 좋습니다. "
            "이후 배경 색상 대비를 조정하면 상품 주목도와 브랜드 메시지 전달력을 함께 높일 수 있습니다."
        ),
        "zone_evaluation_summary": (
            "VP 영역으로서 중앙 착장과 주변 소품이 시즌 분위기를 형성하고 있으며, 대표 상품을 보여주는 기본 역할은 수행하고 있습니다. "
            "다만 보조 상품과 집기의 위계가 약해 첫 시선이 여러 곳으로 분산됩니다. "
            "조명과 간격을 조정하면 핵심 상품의 가시성과 브랜드 메시지 전달력을 더 높일 수 있습니다."
        ),
        "priority_action_summary": (
            "가장 먼저 할 조치는 중앙 착장 주변의 방해 요소를 줄이고 보조 상품의 높이와 간격을 재정렬하는 것입니다. "
            "이 조치가 선행되면 고객 시선이 대표 룩에 먼저 모이고, 이후 연관 상품으로 자연스럽게 이동할 수 있습니다."
        ),
        "detected_issues": [
            "중앙 마네킹 주변 소품과 상품의 간격이 비슷해 주력 상품과 보조 요소의 위계가 약함",
            "우측 하단 의자가 전면 시야와 접근 동선을 일부 방해함",
        ],
        "improvement_actions": [
            "중앙 착장을 기준으로 보조 상품의 높이와 간격을 단계적으로 조정하십시오.",
            "우측 하단 의자를 제거하거나 연출 콘셉트와 연결되는 위치로 옮기십시오.",
            "배경 상품의 색상 수를 줄여 중앙 착장과의 명도 대비를 확보하십시오.",
        ],
        "overall_improvement_summary": (
            "중앙 착장 주변의 간격, 높이, 배경 대비를 정리하면 VP 존의 핵심 상품 가시성이 개선됩니다. "
            "방해 집기를 줄이고 조명 포인트를 조정하면 원거리 주목도와 브랜드 메시지 전달력이 함께 높아집니다."
        ),
        "criteria_evaluations": [
            {
                "criterion": "연출 콘셉트의 시각적 일관성",
                "score": 82,
                "evidence": "색상 계열과 소재감이 반복되어 전체 콘셉트가 비교적 일관됩니다.",
                "issue": "일부 보조 소품의 역할이 불명확해 콘셉트 집중도가 약해질 수 있습니다.",
                "suggestion": "시즌 메시지와 직접 연결되는 소품만 남겨 콘셉트 밀도를 높이십시오.",
            },
            {
                "criterion": "시각적 핵심 상품의 가시성과 강조도",
                "score": 78,
                "evidence": "중앙 착장이 전면에 있어 핵심 상품의 실루엣은 확인됩니다.",
                "issue": "주변 상품과 집기가 비슷한 시각적 비중을 가져 핵심 상품 강조가 약합니다.",
                "suggestion": "조명과 높이 차이를 활용해 중앙 착장을 먼저 보이게 하십시오.",
            },
            {
                "criterion": "포컬 포인트 및 시선 유도",
                "score": 80,
                "evidence": "마네킹 위치가 포컬 포인트를 형성합니다.",
                "issue": "하단 집기와 배경 상품 때문에 시선 흐름이 분산됩니다.",
                "suggestion": "하단 방해 요소를 정리하고 시선이 중앙에서 주변으로 이동하도록 간격을 조정하십시오.",
            },
            {
                "criterion": "구성·비례·여백의 균형",
                "score": 82,
                "evidence": "전면 공간과 배경 공간이 과도하게 비어 있지는 않습니다.",
                "issue": "일부 요소 간 간격이 균등해 위계가 약하게 보입니다.",
                "suggestion": "주력 상품 주변에는 여백을 더 두고 보조 상품은 한 단계 뒤로 배치하십시오.",
            },
            {
                "criterion": "색채·조명·소품의 조화",
                "score": 80,
                "evidence": "색상과 조명은 상품 식별에 충분합니다.",
                "issue": "배경 색상 대비가 약해 중심 착장이 강하게 분리되지 않습니다.",
                "suggestion": "스포트라이트나 배경 명도 차이를 활용해 중심 상품의 질감과 윤곽을 강조하십시오.",
            },
        ],
    }
    return (
        "반드시 아래 예시와 같은 JSON 객체만 반환하세요. 설명 문장, markdown, 코드블록은 쓰지 마세요.\n"
        + json.dumps(example, ensure_ascii=False, indent=2)
    )


def apply_defaults(result: dict[str, Any], user_zone: str) -> dict[str, Any]:
    result.setdefault("user_selected_zone", user_zone)
    result.setdefault("ai_detected_zone", "UNKNOWN")
    result.setdefault("zone_confidence", 0)
    result.setdefault("store_type_assumption", "UNKNOWN")
    result.setdefault(
        "photo_quality",
        {
            "score": 0,
            "is_blurry": False,
            "is_tilted": False,
            "has_background_interference": False,
            "needs_retake": False,
            "comment": "",
        },
    )
    result.setdefault("mannequin", {"exists": False, "type": "none", "has_head": False, "comment": ""})
    result.setdefault("obstacles", [])
    result.setdefault(
        "scores",
        {
            "layout": 0,
            "presentation_mood": 0,
            "brand_fit": 0,
            "color_harmony": 0,
            "cleanliness": 0,
            "customer_attention": 0,
            "season_concept_fit": 0,
        },
    )
    result.setdefault("total_score", 0)
    result.setdefault("grade", grade_from_score(result.get("total_score")))
    result.setdefault("positive_points", [])
    result.setdefault("critical_issues", [])
    result.setdefault("improvement_suggestions", [])
    result.setdefault("final_summary", "")
    result.setdefault("zone_evaluation_summary", result.get("final_summary", ""))
    result.setdefault("priority_action_summary", result.get("final_summary", ""))
    result.setdefault("detected_issues", result.get("critical_issues", []))
    result.setdefault("improvement_actions", result.get("improvement_suggestions", []))
    result.setdefault("overall_improvement_summary", list_to_lines(result.get("improvement_suggestions", [])))
    criteria = result.get("criteria_evaluations")
    if not isinstance(criteria, list):
        criteria = []
    by_name = {str(item.get("criterion", "")).strip(): item for item in criteria if isinstance(item, dict)}
    normalized_criteria = []
    for criterion in criteria_for_zone(user_zone):
        item = by_name.get(criterion, {})
        normalized_criteria.append(
            {
                "criterion": criterion,
                "score": item.get("score"),
                "evidence": item.get("evidence", ""),
                "issue": item.get("issue", ""),
                "suggestion": item.get("suggestion", ""),
            }
        )
    result["criteria_evaluations"] = normalized_criteria
    return result


def analyze_images(images: list[dict[str, Any]], options: dict[str, Any]) -> dict[str, Any]:
    if not images:
        raise ValueError("At least one image is required.")

    requested_model = (options.get("modelName") or DEFAULT_MODEL).strip()
    zone = normalize_zone(options.get("zoneMode"))
    content = [{"type": "text", "text": build_user_text(options, len(images)) + "\n\n" + schema_instruction()}]
    content.extend(image_content_items(images))

    base_payload = {
        "messages": [
            {"role": "system", "content": system_prompt()},
            {"role": "user", "content": content},
        ],
        "temperature": float(options.get("temperature", 0.2) or 0.2),
        "max_tokens": int(options.get("maxTokens", 2200) or 2200),
    }

    errors: list[str] = []
    raw = None
    parsed = None
    model = requested_model
    for candidate_model in lmstudio_model_candidates(requested_model):
        request_variants = [
            {
                **base_payload,
                "model": candidate_model,
                "response_format": {
                    "type": "json_schema",
                    "json_schema": {
                        "name": "vmd_evaluation_result",
                        "strict": True,
                        "schema": vmd_json_schema(),
                    },
                },
            },
            {**base_payload, "model": candidate_model},
        ]
        for payload in request_variants:
            try:
                candidate_raw = lmstudio_request(payload)
            except RuntimeError as exc:
                message = str(exc)
                errors.append(f"{candidate_model}: {message}")
                if "HTTP 400" not in message and not is_retriable_lmstudio_error(message):
                    raise
                continue
            try:
                message = candidate_raw["choices"][0]["message"]
                candidate_parsed = parse_model_content(message.get("content", ""))
            except Exception as exc:
                errors.append(f"{candidate_model}: invalid JSON response: {exc}")
                continue
            raw = candidate_raw
            parsed = candidate_parsed
            model = candidate_model
            break
        if raw is not None:
            break
    if raw is None:
        raise RuntimeError("LM Studio request failed after model and JSON fallbacks: " + " | ".join(errors))

    parsed = apply_defaults(parsed, zone)
    parsed["user_selected_zone"] = zone
    parsed["grade"] = parsed.get("grade") or grade_from_score(parsed.get("total_score"))
    return {"result": parsed, "raw": raw, "model": model}


def list_to_lines(value: Any) -> str:
    if isinstance(value, list):
        lines = []
        for item in value:
            if isinstance(item, dict):
                lines.append("; ".join(f"{k}: {v}" for k, v in item.items()))
            else:
                lines.append(str(item))
        return "\n".join(lines)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else str(value)


def record_zone(record: dict[str, Any]) -> str:
    result = record.get("result", {})
    return normalize_zone(result.get("user_selected_zone"))


def criteria_evaluation_map(result: dict[str, Any]) -> dict[str, dict[str, Any]]:
    evaluations = result.get("criteria_evaluations", [])
    if not isinstance(evaluations, list):
        return {}
    return {
        str(item.get("criterion", "")).strip(): item
        for item in evaluations
        if isinstance(item, dict) and str(item.get("criterion", "")).strip()
    }


def result_to_row(record: dict[str, Any], zone: str | None = None) -> list[Any]:
    result = record.get("result", {})
    photo = result.get("photo_quality", {})
    selected_zone = normalize_zone(zone or result.get("user_selected_zone"))
    row = [
        "",
        result.get("user_selected_zone", ""),
        result.get("ai_detected_zone", ""),
        result.get("total_score", ""),
        photo.get("score", ""),
        photo.get("comment", ""),
        result.get("zone_evaluation_summary", ""),
        result.get("priority_action_summary", result.get("final_summary", "")),
        list_to_lines(result.get("detected_issues", result.get("critical_issues", []))),
        list_to_lines(result.get("improvement_actions", result.get("improvement_suggestions", []))),
        result.get("overall_improvement_summary", list_to_lines(result.get("improvement_suggestions", []))),
        format_elapsed(record.get("elapsed_seconds")),
    ]
    by_criterion = criteria_evaluation_map(result)
    for criterion in criteria_for_zone(selected_zone):
        item = by_criterion.get(criterion, {})
        row.extend(
            [
                item.get("score"),
                item.get("evidence", ""),
                item.get("issue", ""),
                item.get("suggestion", ""),
            ]
        )
    return row


def save_json(name: str, payload: dict[str, Any]) -> str:
    ensure_dirs()
    path = JSON_DIR / f"{safe_file_name(name)}_{timestamp()}.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(path.relative_to(ROOT))


def image_data_url_to_thumbnail(data_url: str, max_size: tuple[int, int] = (120, 90)) -> BytesIO | None:
    if not data_url.startswith("data:image/") or "," not in data_url or PillowImage is None:
        return None
    try:
        encoded = data_url.split(",", 1)[1]
        raw = base64.b64decode(encoded, validate=True)
        image = PillowImage.open(BytesIO(raw))
        if ImageOps is not None:
            image = ImageOps.exif_transpose(image)
        image.thumbnail(max_size)
        if image.mode not in {"RGB", "RGBA"}:
            image = image.convert("RGB")
        stream = BytesIO()
        image.save(stream, format="PNG")
        stream.seek(0)
        return stream
    except Exception:
        return None


def jpeg_exif_orientation(data: bytes) -> int:
    if len(data) < 4 or data[:2] != b"\xff\xd8":
        return 1
    offset = 2
    while offset + 4 <= len(data):
        if data[offset] != 0xFF:
            return 1
        marker = data[offset + 1]
        offset += 2
        if marker in {0xD8, 0xD9}:
            continue
        if offset + 2 > len(data):
            return 1
        segment_length = struct.unpack(">H", data[offset : offset + 2])[0]
        segment = data[offset + 2 : offset + segment_length]
        offset += segment_length
        if marker != 0xE1 or not segment.startswith(b"Exif\x00\x00"):
            continue
        tiff = segment[6:]
        if len(tiff) < 8:
            return 1
        if tiff[:2] == b"II":
            endian = "<"
        elif tiff[:2] == b"MM":
            endian = ">"
        else:
            return 1
        if struct.unpack(endian + "H", tiff[2:4])[0] != 42:
            return 1
        ifd_offset = struct.unpack(endian + "I", tiff[4:8])[0]
        if ifd_offset + 2 > len(tiff):
            return 1
        entry_count = struct.unpack(endian + "H", tiff[ifd_offset : ifd_offset + 2])[0]
        entries_start = ifd_offset + 2
        for index in range(entry_count):
            entry_start = entries_start + index * 12
            entry = tiff[entry_start : entry_start + 12]
            if len(entry) < 12:
                return 1
            tag, value_type, count = struct.unpack(endian + "HHI", entry[:8])
            if tag == 0x0112 and value_type == 3 and count == 1:
                return struct.unpack(endian + "H", entry[8:10])[0]
    return 1


def image_size(data: bytes, extension: str) -> tuple[int, int] | None:
    if extension == "png" and len(data) >= 24 and data[:8] == b"\x89PNG\r\n\x1a\n":
        return struct.unpack(">II", data[16:24])
    if extension == "jpg" and len(data) >= 4 and data[:2] == b"\xff\xd8":
        offset = 2
        while offset + 9 <= len(data):
            if data[offset] != 0xFF:
                return None
            marker = data[offset + 1]
            offset += 2
            if marker in {0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF}:
                segment_length = struct.unpack(">H", data[offset : offset + 2])[0]
                segment = data[offset + 2 : offset + segment_length]
                if len(segment) >= 5:
                    height, width = struct.unpack(">HH", segment[1:5])
                    return width, height
                return None
            if marker in {0xD8, 0xD9}:
                continue
            if offset + 2 > len(data):
                return None
            segment_length = struct.unpack(">H", data[offset : offset + 2])[0]
            offset += segment_length
    return None


def fit_image_extent(data: bytes, extension: str, orientation: int, max_width: int = 120, max_height: int = 90) -> tuple[int, int]:
    size = image_size(data, extension)
    if size is None:
        width, height = max_width, max_height
    else:
        width, height = size
    if orientation in {5, 6, 7, 8}:
        display_width, display_height = height, width
    else:
        display_width, display_height = width, height
    scale = min(max_width / max(1, display_width), max_height / max(1, display_height), 1)
    display_width_px = max(1, round(display_width * scale))
    display_height_px = max(1, round(display_height * scale))
    if orientation in {5, 6, 7, 8}:
        width_px, height_px = display_height_px, display_width_px
    else:
        width_px, height_px = display_width_px, display_height_px
    return width_px * 9525, height_px * 9525


def excel_rotation_from_orientation(orientation: int) -> int:
    rotations = {
        3: 180 * 60000,
        6: 90 * 60000,
        8: 270 * 60000,
    }
    return rotations.get(orientation, 0)


def image_data_url_to_media(data_url: str) -> dict[str, Any] | None:
    if not data_url.startswith("data:image/") or "," not in data_url:
        return None
    header, encoded = data_url.split(",", 1)
    mime = header[5:].split(";", 1)[0].lower()
    if mime == "image/png":
        extension = "png"
    elif mime in {"image/jpeg", "image/jpg"}:
        extension = "jpg"
        mime = "image/jpeg"
    else:
        return None
    try:
        data = base64.b64decode(encoded, validate=True)
    except Exception:
        return None
    orientation = jpeg_exif_orientation(data) if extension == "jpg" else 1
    cx, cy = fit_image_extent(data, extension, orientation)
    return {
        "data": data,
        "extension": extension,
        "content_type": mime,
        "orientation": orientation,
        "rotation": excel_rotation_from_orientation(orientation),
        "cx": cx,
        "cy": cy,
    }


def excel_column_letter(col_idx: int) -> str:
    letters = ""
    while col_idx:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def xlsx_inline_cell(row_idx: int, col_idx: int, value: Any, style: int | None = None) -> str:
    ref = f"{excel_column_letter(col_idx)}{row_idx}"
    style_attr = f' s="{style}"' if style is not None else ""
    raw_text = "" if value is None else str(value)
    cleaned = "".join(
        char
        for char in raw_text
        if char in {"\t", "\n", "\r"} or ord(char) >= 0x20
    )
    text = escape(cleaned)
    return f'<c r="{ref}" t="inlineStr"{style_attr}><is><t xml:space="preserve">{text}</t></is></c>'


def save_excel_pure_python(records: list[dict[str, Any]], prefix: str) -> str:
    path = EXCEL_DIR / f"{prefix}_{timestamp()}.xlsx"
    zone = record_zone(records[0]) if records else "VP"
    headers = excel_headers_for_zone(zone)
    rows = [headers, *[result_to_row(record, zone) for record in records]]
    image_parts = []
    for row_idx, record in enumerate(records, start=2):
        media = image_data_url_to_media(record.get("image_data_url", ""))
        if media is None:
            continue
        image_parts.append(
            {
                "row_idx": row_idx,
                "data": media["data"],
                "extension": media["extension"],
                "content_type": media["content_type"],
                "rotation": media["rotation"],
                "cx": media["cx"],
                "cy": media["cy"],
                "name": f'image{len(image_parts) + 1}.{media["extension"]}',
                "rid": f"rId{len(image_parts) + 1}",
            }
        )

    content_types = [
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
        '<Default Extension="xml" ContentType="application/xml"/>',
    ]
    for extension in sorted({part["extension"] for part in image_parts}):
        content_type = "image/png" if extension == "png" else "image/jpeg"
        content_types.append(f'<Default Extension="{extension}" ContentType="{content_type}"/>')
    content_types.extend(
        [
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>',
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>',
            '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>',
            '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>',
        ]
    )
    if image_parts:
        content_types.append(
            '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
        )

    col_widths = excel_column_widths_for_zone(zone)
    cols_xml = "".join(
        f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>'
        for idx, width in col_widths.items()
    )
    row_xml = []
    for row_idx, values in enumerate(rows, start=1):
        attrs = f' r="{row_idx}"'
        if row_idx == 1:
            attrs += ' ht="24" customHeight="1"'
        if row_idx > 1:
            attrs += ' ht="72" customHeight="1"'
        cells = "".join(xlsx_inline_cell(row_idx, col_idx, value, 1 if row_idx == 1 else None) for col_idx, value in enumerate(values, start=1))
        row_xml.append(f"<row{attrs}>{cells}</row>")
    drawing_ref = '<drawing r:id="rId1"/>' if image_parts else ""
    worksheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
        "</sheetView></sheetViews>"
        f"<cols>{cols_xml}</cols>"
        f'<sheetData>{"".join(row_xml)}</sheetData>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>'
        f"{drawing_ref}"
        "</worksheet>"
    )

    drawing_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
    )
    for idx, part in enumerate(image_parts, start=1):
        zero_based_row = part["row_idx"] - 1
        drawing_xml += (
            "<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>0</xdr:col><xdr:colOff>95250</xdr:colOff><xdr:row>{zero_based_row}</xdr:row><xdr:rowOff>95250</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{part["cx"]}" cy="{part["cy"]}"/>'
            "<xdr:pic>"
            f'<xdr:nvPicPr><xdr:cNvPr id="{idx}" name="Upload Image {idx}"/><xdr:cNvPicPr/></xdr:nvPicPr>'
            f'<xdr:blipFill><a:blip r:embed="{part["rid"]}"/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
            f'<xdr:spPr><a:xfrm rot="{part["rotation"]}"><a:off x="0" y="0"/><a:ext cx="{part["cx"]}" cy="{part["cy"]}"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
            "</xdr:pic><xdr:clientData/>"
            "</xdr:oneCellAnchor>"
        )
    drawing_xml += "</xdr:wsDr>"

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as xlsx:
        xlsx.writestr("[Content_Types].xml", f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">{"".join(content_types)}</Types>')
        xlsx.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
            '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
            "</Relationships>",
        )
        xlsx.writestr(
            "docProps/core.xml",
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><dc:title>VMD 분석 결과</dc:title><dc:creator>AX R&amp;D VMD</dc:creator><dcterms:created xsi:type="dcterms:W3CDTF">{datetime.utcnow().isoformat()}Z</dcterms:created></cp:coreProperties>',
        )
        xlsx.writestr(
            "docProps/app.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"><Application>AX R&amp;D VMD</Application></Properties>',
        )
        xlsx.writestr(
            "xl/workbook.xml",
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="{zone} 결과" sheetId="1" r:id="rId1"/></sheets></workbook>',
        )
        xlsx.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>',
        )
        xlsx.writestr(
            "xl/styles.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><fonts count="2"><font><sz val="11"/><color theme="1"/><name val="Calibri"/></font><font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font></fonts><fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FF1F4E78"/><bgColor indexed="64"/></patternFill></fill></fills><borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders><cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs><cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf><xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf></cellXfs><cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles></styleSheet>',
        )
        xlsx.writestr("xl/worksheets/sheet1.xml", worksheet_xml)
        if image_parts:
            xlsx.writestr(
                "xl/worksheets/_rels/sheet1.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>',
            )
            drawing_rels = "".join(
                f'<Relationship Id="{part["rid"]}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/{part["name"]}"/>'
                for part in image_parts
            )
            xlsx.writestr(
                "xl/drawings/_rels/drawing1.xml.rels",
                f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{drawing_rels}</Relationships>',
            )
            xlsx.writestr("xl/drawings/drawing1.xml", drawing_xml)
            for part in image_parts:
                xlsx.writestr(f'xl/media/{part["name"]}', part["data"])
    return str(path.relative_to(ROOT))


def save_excel(records: list[dict[str, Any]], prefix: str = "vmd_results") -> str:
    ensure_dirs()
    if Workbook is None:
        return save_excel_pure_python(records, prefix)

    zone = record_zone(records[0]) if records else "VP"
    headers = excel_headers_for_zone(zone)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{zone} 결과"
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    image_streams = []
    for record in records:
        row_idx = ws.max_row + 1
        ws.append(result_to_row(record, zone))
        if ExcelImage is not None:
            stream = image_data_url_to_thumbnail(record.get("image_data_url", ""))
            if stream is not None:
                image_streams.append(stream)
                embedded = ExcelImage(stream)
                embedded.anchor = f"A{row_idx}"
                ws.add_image(embedded)
                ws.row_dimensions[row_idx].height = 72

    widths = excel_column_widths_for_zone(zone)
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(col_idx, 14)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    path = EXCEL_DIR / f"{prefix}_{timestamp()}.xlsx"
    wb.save(path)
    return str(path.relative_to(ROOT))


def make_record(
    image_names: str,
    analysis: dict[str, Any],
    status: str = "success",
    error: str = "",
    image_data_url: str = "",
    elapsed_seconds: float | None = None,
) -> dict[str, Any]:
    json_path = save_json(safe_file_name(image_names[:80]), analysis)
    return {
        "image_names": image_names,
        "image_data_url": image_data_url,
        "result": analysis.get("result", {}),
        "raw": analysis.get("raw", {}),
        "json_path": json_path,
        "status": status,
        "error": error,
        "elapsed_seconds": elapsed_seconds,
    }


PDF_FONT_REGULAR = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"


def setup_pdf_fonts() -> None:
    global PDF_FONT_REGULAR, PDF_FONT_BOLD
    if not REPORTLAB_AVAILABLE:
        return
    regular_path = Path("C:/Windows/Fonts/malgun.ttf")
    bold_path = Path("C:/Windows/Fonts/malgunbd.ttf")
    if not regular_path.exists() or not bold_path.exists():
        return
    if "VmdMalgun" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VmdMalgun", str(regular_path)))
    if "VmdMalgunBold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("VmdMalgunBold", str(bold_path)))
    PDF_FONT_REGULAR = "VmdMalgun"
    PDF_FONT_BOLD = "VmdMalgunBold"


def pdf_markup(value: Any, fallback: str = "-") -> str:
    text = list_to_lines(value).strip() if value is not None else ""
    return html.escape(text or fallback).replace("\n", "<br/>")


def pdf_image_for_record(record: dict[str, Any], max_width: float = 220, max_height: float = 150) -> Any:
    media = image_data_url_to_media(record.get("image_data_url", ""))
    if media is None:
        return Paragraph("대표 이미지가 없습니다.", ParagraphStyle("missing-image", fontName=PDF_FONT_REGULAR, fontSize=9, textColor=pdf_colors.HexColor("#6F6A61")))
    stream = BytesIO(media["data"])
    width, height = ImageReader(stream).getSize()
    scale = min(max_width / max(1, width), max_height / max(1, height), 1)
    image = ReportLabImage(stream, width=max(1, width * scale), height=max(1, height * scale))
    image.hAlign = "CENTER"
    return image


def draw_pdf_footer(canvas: Any, doc: Any) -> None:
    canvas.saveState()
    canvas.setFillColor(pdf_colors.HexColor("#F7F5F0"))
    canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
    canvas.setFillColor(pdf_colors.HexColor("#B79B61"))
    canvas.rect(doc.leftMargin, A4[1] - 11 * mm, A4[0] - doc.leftMargin - doc.rightMargin, 1.2, fill=1, stroke=0)
    canvas.setStrokeColor(pdf_colors.HexColor("#DED8CE"))
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, 18 * mm, A4[0] - doc.rightMargin, 18 * mm)
    canvas.setFont(PDF_FONT_REGULAR, 7.5)
    canvas.setFillColor(pdf_colors.HexColor("#6F6A61"))
    canvas.drawString(doc.leftMargin, 11 * mm, "AX R&D VMD - 분석 결과 보고서")
    canvas.drawRightString(A4[0] - doc.rightMargin, 11 * mm, f"{doc.page}")
    canvas.restoreState()


def save_pdf(records: list[dict[str, Any]], prefix: str = "vmd_results") -> str:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("PDF 저장을 위해 reportlab 패키지가 필요합니다. requirements.txt를 설치한 뒤 다시 시도해 주세요.")
    ensure_dirs()
    setup_pdf_fonts()
    path = PDF_DIR / f"{prefix}_{timestamp()}.pdf"
    styles = getSampleStyleSheet()
    eyebrow_style = ParagraphStyle(
        "pdf-eyebrow",
        parent=styles["Normal"],
        fontName=PDF_FONT_BOLD,
        fontSize=7.5,
        leading=10,
        textColor=pdf_colors.HexColor("#B79B61"),
        spaceAfter=5,
    )
    title_style = ParagraphStyle(
        "pdf-title",
        parent=styles["Title"],
        fontName=PDF_FONT_BOLD,
        fontSize=20,
        leading=26,
        textColor=pdf_colors.HexColor("#111111"),
        alignment=TA_LEFT,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "pdf-subtitle",
        parent=styles["Normal"],
        fontName=PDF_FONT_REGULAR,
        fontSize=9,
        leading=13,
        textColor=pdf_colors.HexColor("#6F6A61"),
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "pdf-section",
        parent=styles["Heading2"],
        fontName=PDF_FONT_BOLD,
        fontSize=12,
        leading=16,
        textColor=pdf_colors.HexColor("#111111"),
        spaceBefore=10,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "pdf-body",
        parent=styles["BodyText"],
        fontName=PDF_FONT_REGULAR,
        fontSize=8.8,
        leading=13,
        textColor=pdf_colors.HexColor("#3D3932"),
    )
    small_style = ParagraphStyle(
        "pdf-small",
        parent=body_style,
        fontSize=7.6,
        leading=10.5,
    )
    label_style = ParagraphStyle(
        "pdf-label",
        parent=body_style,
        fontName=PDF_FONT_BOLD,
        fontSize=7.8,
        leading=10,
        textColor=pdf_colors.HexColor("#6F6A61"),
    )
    issue_header_dark_style = ParagraphStyle(
        "pdf-issue-header-dark",
        parent=section_style,
        textColor=pdf_colors.white,
        spaceBefore=0,
        spaceAfter=0,
    )
    issue_header_gold_style = ParagraphStyle(
        "pdf-issue-header-gold",
        parent=section_style,
        textColor=pdf_colors.HexColor("#111111"),
        spaceBefore=0,
        spaceAfter=0,
    )
    table_header_style = ParagraphStyle(
        "pdf-table-header",
        parent=body_style,
        fontName=PDF_FONT_BOLD,
        fontSize=7.5,
        leading=9.5,
        textColor=pdf_colors.white,
        alignment=TA_CENTER,
    )
    table_cell_style = ParagraphStyle(
        "pdf-table-cell",
        parent=body_style,
        fontSize=7.3,
        leading=9.5,
    )
    score_style = ParagraphStyle(
        "pdf-score",
        parent=table_cell_style,
        fontName=PDF_FONT_BOLD,
        fontSize=12,
        leading=14,
        textColor=pdf_colors.HexColor("#111111"),
        alignment=TA_CENTER,
    )

    story: list[Any] = []
    for record_index, record in enumerate(records):
        result = record.get("result", {})
        zone = record_zone(record)
        photo = result.get("photo_quality", {})
        mannequin = result.get("mannequin", {})
        title = f"VMD 이미지 분석 리포트 - {zone}"
        image_names = record.get("image_names", "분석 이미지")
        detected_zone = result.get("ai_detected_zone", "UNKNOWN")
        confidence = round(float(result.get("zone_confidence", 0) or 0) * 100)

        story.append(Paragraph("AX R&D VISUAL MERCHANDISING", eyebrow_style))
        story.append(Paragraph(html.escape(title), title_style))
        story.append(Paragraph(f"분석 이미지: {html.escape(str(image_names))}<br/>분석 시간: {html.escape(format_elapsed(record.get('elapsed_seconds')))}", subtitle_style))

        meta_data = [
            [Paragraph("사용자 지정 구역", label_style), Paragraph(pdf_markup(result.get("user_selected_zone", zone)), body_style), Paragraph("AI 감지 구역", label_style), Paragraph(pdf_markup(detected_zone), body_style)],
            [Paragraph("구역 신뢰도", label_style), Paragraph(f"{confidence}%", body_style), Paragraph("총점 / 등급", label_style), Paragraph(f"{pdf_markup(result.get('total_score'))} / {pdf_markup(result.get('grade'))}", body_style)],
            [Paragraph("사진 품질", label_style), Paragraph(f"{pdf_markup(photo.get('score'))}점 - {pdf_markup(photo.get('comment'))}", small_style), Paragraph("마네킹", label_style), Paragraph(f"{('있음' if mannequin.get('exists') else '없음')} - {pdf_markup(mannequin.get('type'))}", small_style)],
        ]
        meta_table = Table(meta_data, colWidths=[80, 170, 70, 185], hAlign="LEFT")
        meta_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), pdf_colors.white),
            ("BOX", (0, 0), (-1, -1), 0.6, pdf_colors.HexColor("#DED8CE")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, pdf_colors.HexColor("#DED8CE")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 10))

        summary_data = [
            [
                [Paragraph("영역별 종합 평가", section_style), Paragraph(pdf_markup(result.get("zone_evaluation_summary", result.get("final_summary"))), body_style)],
                [Paragraph("우선 개선 방향", section_style), Paragraph(pdf_markup(result.get("priority_action_summary", result.get("final_summary"))), body_style)],
            ],
            [pdf_image_for_record(record), Paragraph(f"<b>최종 요약</b><br/>{pdf_markup(result.get('overall_improvement_summary', result.get('final_summary')))}", body_style)],
        ]
        summary_table = Table(summary_data, colWidths=[252, 253], hAlign="LEFT")
        summary_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOX", (0, 0), (-1, -1), 0.6, pdf_colors.HexColor("#DED8CE")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, pdf_colors.HexColor("#DED8CE")),
            ("BACKGROUND", (0, 0), (-1, 0), pdf_colors.HexColor("#EADFCA")),
            ("BACKGROUND", (0, 1), (-1, 1), pdf_colors.white),
            ("LINEABOVE", (0, 0), (-1, 0), 1.2, pdf_colors.HexColor("#B79B61")),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(summary_table)
        story.append(Paragraph("항목별 평가", section_style))

        by_criterion = criteria_evaluation_map(result)
        criteria_rows = [[
            Paragraph("평가항목", table_header_style),
            Paragraph("점수", table_header_style),
            Paragraph("근거", table_header_style),
            Paragraph("문제점", table_header_style),
            Paragraph("개선안", table_header_style),
        ]]
        for criterion in criteria_for_zone(zone):
            item = by_criterion.get(criterion, {})
            criteria_rows.append([
                Paragraph(pdf_markup(criterion), table_cell_style),
                Paragraph(pdf_markup(item.get("score")), score_style),
                Paragraph(pdf_markup(item.get("evidence")), table_cell_style),
                Paragraph(pdf_markup(item.get("issue")), table_cell_style),
                Paragraph(pdf_markup(item.get("suggestion")), table_cell_style),
            ])
        criteria_table = LongTable(criteria_rows, colWidths=[105, 38, 120, 120, 120], repeatRows=1, hAlign="LEFT")
        criteria_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), pdf_colors.HexColor("#050505")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [pdf_colors.white, pdf_colors.HexColor("#F7F5F0")]),
            ("LINEBELOW", (0, 0), (-1, 0), 1.2, pdf_colors.HexColor("#B79B61")),
            ("BOX", (0, 0), (-1, -1), 0.6, pdf_colors.HexColor("#DED8CE")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, pdf_colors.HexColor("#DED8CE")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(criteria_table)
        story.append(Spacer(1, 8))

        issue_data = [[
            Paragraph("감지된 문제점", issue_header_dark_style),
            Paragraph("개선 방향", issue_header_gold_style),
        ], [
            Paragraph(pdf_markup(result.get("detected_issues", result.get("critical_issues", []))), body_style),
            Paragraph(pdf_markup(result.get("improvement_actions", result.get("improvement_suggestions", []))), body_style),
        ]]
        issue_table = Table(issue_data, colWidths=[252, 253], hAlign="LEFT")
        issue_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), pdf_colors.HexColor("#050505")),
            ("BACKGROUND", (1, 0), (1, 0), pdf_colors.HexColor("#EADFCA")),
            ("LINEABOVE", (0, 0), (-1, 0), 1.2, pdf_colors.HexColor("#B79B61")),
            ("BOX", (0, 0), (-1, -1), 0.6, pdf_colors.HexColor("#DED8CE")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, pdf_colors.HexColor("#DED8CE")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(issue_table)
        if record_index < len(records) - 1:
            story.append(PageBreak())

    document = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=16 * mm,
        bottomMargin=25 * mm,
        title="VMD 이미지 분석 결과",
        author="AX R&D VMD",
    )
    document.build(story, onFirstPage=draw_pdf_footer, onLaterPages=draw_pdf_footer)
    return str(path.relative_to(ROOT))


def json_response(handler: BaseHTTPRequestHandler, status: int, payload: dict[str, Any]) -> None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def read_json_body(handler: BaseHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length", "0"))
    body = handler.rfile.read(length)
    if not body:
        return {}
    return json.loads(body.decode("utf-8"))


def content_type_for(path: Path) -> str:
    suffix = path.suffix.lower()
    return {
        ".html": "text/html; charset=utf-8",
        ".css": "text/css; charset=utf-8",
        ".js": "application/javascript; charset=utf-8",
        ".json": "application/json; charset=utf-8",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".pdf": "application/pdf",
        ".csv": "text/csv; charset=utf-8",
    }.get(suffix, "application/octet-stream")


def friendly_error_message(exc: Exception) -> str:
    message = str(exc)
    if "Failed to load model" in message or "exited before becoming healthy" in message:
        return (
            "LM Studio에서 Gemma 12B 계열 모델 로드에 실패했습니다. "
            "LM Studio에서 모델을 언로드한 뒤 다시 로드하고, Local Server를 재시작한 다음 다시 분석해 주세요."
        )
    if "Connection refused" in message or "연결을 거부" in message or "WinError 10061" in message:
        return "LM Studio 서버에 연결할 수 없습니다. LM Studio에서 Gemma 4 12B 모델을 로드하고 Local Server를 켜주세요."
    if "LM Studio HTTP 400" in message:
        return (
            "LM Studio가 분석 요청을 거절했습니다. 모델명이 정확한지, 현재 모델이 이미지 입력을 지원하는지 확인해 주세요. "
        )
    return message


class VmdHandler(BaseHTTPRequestHandler):
    server_version = "VMDWebApp/0.1"

    def log_message(self, fmt: str, *args: Any) -> None:
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/health":
            ok, message = check_lmstudio()
            json_response(
                self,
                200,
                {
                    "ok": True,
                    "lmstudio": ok,
                    "message": message,
                    "baseUrl": LMSTUDIO_BASE_URL,
                    "defaultModel": DEFAULT_MODEL,
                    "fallbackModels": lmstudio_model_candidates(DEFAULT_MODEL)[1:],
                },
            )
            return

        if parsed.path == "/api/download":
            params = urllib.parse.parse_qs(parsed.query)
            rel = params.get("file", [""])[0]
            self.serve_download(rel)
            return

        rel_path = "index.html" if parsed.path == "/" else parsed.path.lstrip("/")
        path = STATIC_DIR / rel_path
        if not path.exists() or not path.is_file():
            self.send_error(404, "Not found")
            return
        self.serve_file(path)

    def do_POST(self) -> None:
        try:
            if self.path == "/api/analyze":
                self.handle_analyze()
                return
            if self.path == "/api/batch-analyze":
                self.handle_batch()
                return
            self.send_error(404, "Not found")
        except Exception as exc:
            traceback.print_exc()
            json_response(self, 500, {"ok": False, "error": friendly_error_message(exc), "detail": str(exc)})

    def handle_analyze(self) -> None:
        body = read_json_body(self)
        images = body.get("images", [])
        options = body.get("options", {})
        started = time.time()
        analysis = analyze_images(images, options)
        elapsed_seconds = time.time() - started
        image_names = ", ".join(safe_file_name(img.get("name", "image")) for img in images)
        image_data_url = images[0].get("dataUrl", "") if images else ""
        record = make_record(image_names, analysis, image_data_url=image_data_url, elapsed_seconds=elapsed_seconds)
        result_prefix = f"{record_zone(record).lower()}_results"
        excel_path = save_excel([record], prefix=result_prefix)
        pdf_path = save_pdf([record], prefix=result_prefix)
        json_response(
            self,
            200,
            {
                "ok": True,
                "result": analysis["result"],
                "jsonPath": record["json_path"],
                "excelPath": excel_path,
                "downloadUrl": f"/api/download?file={urllib.parse.quote(excel_path)}",
                "pdfPath": pdf_path,
                "pdfDownloadUrl": f"/api/download?file={urllib.parse.quote(pdf_path)}",
                "elapsedSeconds": round(elapsed_seconds, 2),
            },
        )

    def handle_batch(self) -> None:
        body = read_json_body(self)
        images = body.get("images", [])
        options = body.get("options", {})
        records = []
        batch_results = []
        started = time.time()
        for image in images:
            name = safe_file_name(image.get("name", "image"))
            original_name = image.get("name", "image")
            try:
                image_started = time.time()
                analysis = analyze_images([image], options)
                record = make_record(
                    name,
                    analysis,
                    image_data_url=image.get("dataUrl", ""),
                    elapsed_seconds=time.time() - image_started,
                )
                records.append(record)
                batch_results.append(
                    {
                        "imageName": original_name,
                        "result": record["result"],
                        "jsonPath": record["json_path"],
                        "status": record["status"],
                    }
                )
            except Exception as exc:
                friendly = friendly_error_message(exc)
                error_result = {
                    "result": {
                        "user_selected_zone": normalize_zone(options.get("zoneMode")),
                        "ai_detected_zone": "UNKNOWN",
                        "zone_confidence": 0,
                        "store_type_assumption": options.get("storeType", "UNKNOWN"),
                        "photo_quality": {
                            "score": 0,
                            "is_blurry": False,
                            "is_tilted": False,
                            "has_background_interference": False,
                            "needs_retake": False,
                            "comment": "",
                        },
                        "mannequin": {"exists": False, "type": "unknown", "has_head": False, "comment": ""},
                        "obstacles": [],
                        "scores": {
                            "layout": 0,
                            "presentation_mood": 0,
                            "brand_fit": 0,
                            "color_harmony": 0,
                            "cleanliness": 0,
                            "customer_attention": 0,
                            "season_concept_fit": 0,
                        },
                        "total_score": 0,
                        "grade": "분석 실패",
                        "positive_points": [],
                        "critical_issues": [],
                        "improvement_suggestions": [],
                        "final_summary": "",
                    },
                    "raw": {"error": friendly, "original_error": str(exc)},
                }
                record = make_record(
                    name,
                    error_result,
                    status="error",
                    error=friendly,
                    image_data_url=image.get("dataUrl", ""),
                    elapsed_seconds=None,
                )
                records.append(record)
                batch_results.append(
                    {
                        "imageName": original_name,
                        "result": record["result"],
                        "jsonPath": record["json_path"],
                        "status": record["status"],
                        "error": record["error"],
                    }
                )
        zone = normalize_zone(options.get("zoneMode"))
        result_prefix = f"{zone.lower()}_results"
        excel_path = save_excel(records, prefix=result_prefix)
        pdf_path = save_pdf(records, prefix=result_prefix)
        json_response(
            self,
            200,
            {
                "ok": True,
                "count": len(records),
                "results": batch_results,
                "excelPath": excel_path,
                "downloadUrl": f"/api/download?file={urllib.parse.quote(excel_path)}",
                "pdfPath": pdf_path,
                "pdfDownloadUrl": f"/api/download?file={urllib.parse.quote(pdf_path)}",
                "elapsedSeconds": round(time.time() - started, 2),
            },
        )

    def serve_file(self, path: Path) -> None:
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type_for(path))
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def serve_download(self, rel: str) -> None:
        target = (ROOT / rel).resolve()
        try:
            target.relative_to(OUTPUT_DIR.resolve())
        except ValueError:
            self.send_error(403, "Forbidden")
            return
        if not target.exists() or not target.is_file():
            self.send_error(404, "Not found")
            return
        self.send_response(200)
        self.send_header("Content-Type", content_type_for(target))
        self.send_header("Content-Length", str(target.stat().st_size))
        self.send_header("Content-Disposition", f'attachment; filename="{target.name}"')
        self.end_headers()
        self.wfile.write(target.read_bytes())


def run(host: str = "127.0.0.1", port: int = 8000) -> None:
    ensure_dirs()
    server = ThreadingHTTPServer((host, port), VmdHandler)
    print(f"VMD web app running at http://{host}:{port}")
    print(f"LM Studio endpoint: {LMSTUDIO_BASE_URL}")
    print(f"Default model: {DEFAULT_MODEL}")
    server.serve_forever()


if __name__ == "__main__":
    selected_port = int(os.environ.get("PORT", "8000"))
    run(port=selected_port)
